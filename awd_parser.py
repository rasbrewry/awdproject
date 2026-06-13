"""
AWD Field Data Parser
Handles the real-world quirks of Green Carbon water monitoring workbooks:
- Formula strings stored as text ('=22.5-15')
- Multi-row pipe blocks (Monitoring / Dry count / Picture / Time / Climate / Remarks)
- Merged headers, banner rows
- Malformed coordinates ('125.150.100')
- KMZ as ZIP -> doc.kml, namespace-agnostic KML parsing
- Date-indexed AND DAT-indexed monitoring sheets
"""

import re
import io
import zipfile
import datetime
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from shapely.geometry import Point, Polygon

# ---------------------------------------------------------------- value parsing

FORMULA_RE = re.compile(r"^=\s*(-?\d+(?:\.\d+)?)\s*([-+])\s*(-?\d+(?:\.\d+)?)\s*$")
NUMBER_RE = re.compile(r"^-?\d+(?:\.\d+)?$")


def parse_water_value(cell):
    """Convert a monitoring cell into a float, or None.

    Handles:  8.0  |  '=22.5-15'  |  '=0-15'  |  '-15'  |  None  |  'completed'
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return float(cell)
    if isinstance(cell, str):
        s = cell.strip()
        m = FORMULA_RE.match(s)
        if m:
            a, op, b = float(m.group(1)), m.group(2), float(m.group(3))
            return a - b if op == "-" else a + b
        if NUMBER_RE.match(s):
            return float(s)
    return None  # text remarks, 'completed', etc.


def clean_coordinate(raw):
    """Fix malformed coords like '125.150.100' -> 125.150100 (Kahaponan quirk)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    parts = s.split(".")
    if len(parts) > 2:
        s = parts[0] + "." + "".join(parts[1:])
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------- pipe metadata


def read_pipe_info(path_or_stream):
    """Read pipe metadata (number, category, farmer, lat/lon) from the workbook.

    Strategy: scan every sheet for a header row containing 'Pipe No.' plus
    Latitude/Longitude columns. Works with the banner row on top (skiprows=1
    pattern) or without it.
    """
    wb = load_workbook(path_or_stream, read_only=True, data_only=False)
    result = {"pipes": [], "sheet": None, "warnings": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_idx, colmap = None, {}
        rows = []
        for i, row in enumerate(ws.iter_rows(max_row=300, values_only=True)):
            rows.append(row)
            if i > 10 and header_idx is None:
                break
            if header_idx is None and row:
                labels = {str(c).strip().lower(): j for j, c in enumerate(row) if c}
                if any("pipe no" in k for k in labels) and any("latitude" in k for k in labels):
                    header_idx = i
                    for key, j in labels.items():
                        if "pipe no" in key:
                            colmap["pipe"] = j
                        elif "category" in key:
                            colmap["category"] = j
                        elif "farmer" in key:
                            colmap["farmer"] = j
                        elif "latitude" in key:
                            colmap["lat"] = j
                        elif "longitude" in key:
                            colmap["lon"] = j
                        elif "area name" in key and "area" not in colmap:
                            colmap["area"] = j
        if header_idx is None:
            continue

        # stream the remaining rows of this sheet
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            if not row:
                continue
            raw_pipe = row[colmap["pipe"]] if colmap.get("pipe") is not None else None
            if raw_pipe is None:
                continue
            try:
                pipe_no = int(float(raw_pipe))
            except (TypeError, ValueError):
                continue
            lat = clean_coordinate(row[colmap["lat"]]) if "lat" in colmap else None
            lon = clean_coordinate(row[colmap["lon"]]) if "lon" in colmap else None
            cat = row[colmap["category"]] if "category" in colmap else None
            cat = str(cat).strip().upper() if cat else None
            farmer = row[colmap["farmer"]] if "farmer" in colmap else None
            farmer = str(farmer).strip() if farmer else None
            pipe = {
                "pipe": pipe_no,
                "category": cat,
                "farmer": farmer,
                "lat": lat,
                "lon": lon,
            }
            result["pipes"].append(pipe)
            if lat is None or lon is None:
                result["warnings"].append(f"Pipe {pipe_no}: missing coordinates")
        result["sheet"] = sheet_name
        break

    wb.close()
    if not result["pipes"]:
        result["warnings"].append(
            "No pipe metadata found. Expected a sheet with 'Pipe No.', 'Latitude', 'Longitude' columns."
        )
    return result


# ---------------------------------------------------------------- monitoring data


def read_monitoring_sheets(path_or_stream):
    """Find every monitoring sheet in the workbook and extract per-pipe series.

    Returns {sheet_name: {"axis": "date"|"dat", "labels": [...], "series": {pipe: [...]}}}

    Detection logic:
    - header row = first row with >=5 datetimes (date axis) OR a 'DAT' label
      followed by >=5 numbers (DAT axis)
    - pipe data row = numeric col A + 'Monitoring' in col B (date sheets),
      or numeric col A directly under a DAT header (gas-area sheets)
    """
    wb = load_workbook(path_or_stream, read_only=True, data_only=False)
    sheets = {}

    for sheet_name in wb.sheetnames:
        if "read" in sheet_name.lower() and "sample" in sheet_name.lower():
            continue  # instruction sheet
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        axis, header_idx, start_col, labels = None, None, None, []

        for i, row in enumerate(rows[:15]):
            if not row:
                continue
            dt_cols = [j for j, c in enumerate(row) if isinstance(c, datetime.datetime)]
            if len(dt_cols) >= 5:
                axis, header_idx, start_col = "date", i, dt_cols[0]
                labels = [
                    c.strftime("%Y-%m-%d") if isinstance(c, datetime.datetime) else None
                    for c in row[start_col:]
                ]
                break
            first = str(row[0]).strip().upper() if row[0] else ""
            num_cols = [j for j, c in enumerate(row[1:], 1) if isinstance(c, (int, float))]
            if first == "DAT" and len(num_cols) >= 5:
                axis, header_idx, start_col = "dat", i, num_cols[0]
                labels = [
                    str(int(c)) if isinstance(c, (int, float)) else None
                    for c in row[start_col:]
                ]
                break
        if axis is None:
            continue

        series = {}
        for row in rows[header_idx + 1 :]:
            if not row or row[0] is None:
                continue
            try:
                pipe_no = int(float(row[0]))
            except (TypeError, ValueError):
                continue
            # date sheets carry a status label in col B; require 'Monitoring'
            if axis == "date":
                status = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
                if "monitoring" not in status:
                    continue
            values = [
                parse_water_value(row[j]) if j < len(row) else None
                for j in range(start_col, start_col + len(labels))
            ]
            if any(v is not None for v in values):
                series[pipe_no] = values

        if series:
            # trim trailing all-None columns
            last = max(
                (max((j for j, v in enumerate(vals) if v is not None), default=-1))
                for vals in series.values()
            )
            labels = labels[: last + 1]
            series = {p: v[: last + 1] for p, v in series.items()}
            sheets[sheet_name] = {"axis": axis, "labels": labels, "series": series}

    wb.close()
    return sheets


# ---------------------------------------------------------------- KML / KMZ


def _local(tag):
    return tag.rsplit("}", 1)[-1]


def _parse_coord_string(text):
    coords = []
    for token in (text or "").split():
        parts = token.split(",")
        if len(parts) >= 2:
            try:
                coords.append((float(parts[0]), float(parts[1])))  # lon, lat
            except ValueError:
                pass
    return coords


def read_kml_geometries(file_bytes, filename=""):
    """Accepts KMZ or KML bytes. Returns {"polygons":[{name,coords}], "points":[{name,lon,lat}]}."""
    data = file_bytes
    if filename.lower().endswith(".kmz") or data[:2] == b"PK":
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            kml_names = [n for n in z.namelist() if n.lower().endswith(".kml")]
            if not kml_names:
                raise ValueError("KMZ archive contains no .kml file")
            # prefer doc.kml, else first kml
            target = next((n for n in kml_names if n.lower().endswith("doc.kml")), kml_names[0])
            data = z.read(target)

    root = ET.fromstring(data)
    polygons, points = [], []

    def walk(el, folder=""):
        for child in el:
            tag = _local(child.tag)
            if tag == "Folder":
                name_el = next((c for c in child if _local(c.tag) == "Name" or _local(c.tag) == "name"), None)
                walk(child, name_el.text.strip() if name_el is not None and name_el.text else folder)
            elif tag == "Placemark":
                pname = folder
                for c in child:
                    if _local(c.tag) == "name" and c.text:
                        pname = (folder + " / " if folder else "") + c.text.strip()
                for geom in child.iter():
                    gtag = _local(geom.tag)
                    if gtag == "Polygon":
                        for coords_el in geom.iter():
                            if _local(coords_el.tag) == "coordinates":
                                pts = _parse_coord_string(coords_el.text)
                                if len(pts) >= 3:
                                    polygons.append({"name": pname, "coords": pts})
                                break
                    elif gtag == "Point":
                        for coords_el in geom:
                            if _local(coords_el.tag) == "coordinates":
                                pts = _parse_coord_string(coords_el.text)
                                if pts:
                                    points.append({"name": pname, "lon": pts[0][0], "lat": pts[0][1]})
            else:
                walk(child, folder)

    walk(root)
    return {"polygons": polygons, "points": points}


# ---------------------------------------------------------------- cross-check

M_PER_DEG = 111_000  # approx at equator; fine for proximity flags at these latitudes


def cross_check(pipes, geoms, category="AWD", near_threshold_m=15):
    """For each pipe (filtered by category) decide inside / near / outside vs chamber polygons."""
    polys = [(g["name"], Polygon(g["coords"])) for g in geoms["polygons"]]
    results = []
    for p in pipes:
        if category != "ALL" and (p["category"] or "") != category:
            continue
        if p["lat"] is None or p["lon"] is None:
            results.append({**p, "status": "no coordinates", "chamber": None, "distance_m": None})
            continue
        pt = Point(p["lon"], p["lat"])  # KML order: lon, lat
        inside = next((name for name, poly in polys if poly.contains(pt)), None)
        if inside:
            results.append({**p, "status": "inside", "chamber": inside, "distance_m": 0.0})
            continue
        best_name, best_d = None, None
        for name, poly in polys:
            d = poly.exterior.distance(pt) * M_PER_DEG
            if best_d is None or d < best_d:
                best_name, best_d = name, d
        if best_d is not None and best_d <= near_threshold_m:
            results.append({**p, "status": "near", "chamber": best_name, "distance_m": round(best_d, 1)})
        else:
            results.append(
                {**p, "status": "outside", "chamber": best_name, "distance_m": round(best_d, 1) if best_d is not None else None}
            )
    return results
